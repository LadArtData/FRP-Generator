"""Text extraction and section-aware chunking for DOCX, PDF, and XLSX.

Three rules, in priority order:

  1. A chunk never crosses a section boundary.
  2. A chunk never splits a sentence.
  3. A chunk carries the heading it came from.

The chunks in the original library obeyed none of these. Chunk 3 of the
St. Petersburg proposal ended one cover letter and began the next; chunk 1 was
a table-of-contents line. That is why section tags had to be inferred from body
text and smoothed across neighbours in the first place: the boundaries were
wrong, so the labels had nothing solid to attach to. Fixing the boundaries takes
most of the tagging problem with it. Re-chunked, the anchor document went from
thirteen chunks starting mid-sentence to none.

Page furniture is stripped before chunking. Address blocks, phone numbers,
table-of-contents entries and running headers are not answer content, and
embedding them means a requirement about contact procedures retrieves iteria's
letterhead.

Input is bytes rather than a path because documents arrive as uploads.
"""
from __future__ import annotations

import io
import logging
import re
from collections import Counter

from docx import Document

from . import classifier, tagger
from .errors import ValidationFailed
from .vocabulary import DEFAULT_SECTION

log = logging.getLogger("harald.chunking")

# Tokens are estimated as characters / 4. Not a tokenizer count.
TARGET_TOKENS = 300      # aim
MAX_TOKENS = 550         # ceiling before a sentence is forced out alone
MIN_TOKENS = 40          # below this a chunk is merged forward, not kept
OVERLAP_SENTENCES = 1    # carried from the previous chunk for continuity


def est_tokens(text: str) -> int:
    return max(1, round(len(text) / 4))


# ---------------------------------------------------------------------------
# Sentence splitting. Abbreviations are the whole problem: a naive split on
# ". " breaks "St. Petersburg", "iteria.us", "Mr. Poceous", and every decimal
# in a cost table.
# ---------------------------------------------------------------------------
_ABBREV = (
    r"Mr|Mrs|Ms|Dr|Prof|Sr|Jr|St|Ste|Ave|Blvd|Rd|Dept|Est|Inc|Corp|Co|Ltd|LLC|"
    r"No|Nos|Fig|vs|etc|approx|Sect|Sec|Art|Para|p|pp|Vol|Ch|Ref|Att|Attn|"
    r"U\.S|U\.K|D\.C|Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec"
)
_PROTECT = [
    (re.compile(r"\b(" + _ABBREV + r")\.", re.I), r"\1<PD>"),
    (re.compile(r"\b([A-Z])\.(?=\s*[A-Z])"), r"\1<PD>"),            # initials
    (re.compile(r"(\d)\.(\d)"), r"\1<PD>\2"),                        # decimals
    (re.compile(r"\b([a-z]+)\.(us|com|org|gov|net)\b", re.I), r"\1<PD>\2"),
]
_SPLIT = re.compile(r"(?<=[.!?])[\"')\]]*\s+(?=[A-Z\"'(\[])")


def split_sentences(text: str) -> list[str]:
    protected = text
    for pattern, replacement in _PROTECT:
        protected = pattern.sub(replacement, protected)
    out = []
    for part in _SPLIT.split(protected):
        part = part.replace("<PD>", ".").strip()
        if part:
            out.append(part)
    return out


# ---------------------------------------------------------------------------
# Page furniture
# ---------------------------------------------------------------------------
_TOC_LINE = re.compile(r"\.{4,}\s*\d+\s*$|\s\.\s\.\s\.")
_PAGE_NO = re.compile(r"^\s*(page\s+)?\d{1,3}\s*(of\s+\d{1,3})?\s*$", re.I)
_PHONE = re.compile(r"^\s*[\d\-\(\)\.\s]{7,20}\s*$")
_EMAIL_ONLY = re.compile(r"^\s*\S+@\S+\.\S+\s*$")
_URL_ONLY = re.compile(r"^\s*(https?://|www\.)\S+\s*$", re.I)
_ADDRESS = re.compile(
    r"^\s*\d+\s+[\w\s\.]+,?\s+(suite|ste|floor|fl)\b|"
    r"^\s*[\w\s]+,\s*[A-Z]{2}\s+\d{5}(-\d{4})?\s*$", re.I)
_TOC_HEADER = re.compile(r"^\s*table of contents\s*$", re.I)

# Drafting annotations left in a document by whoever wrote it: "[VERIFY RELEASE
# VERSION]", "[CLIENT-SPECIFIC - VERIFY]", "[Contact TBD]". They are not answer
# content, and the anchor document carries 27 of them. Left in the chunk text
# they become voice reference, and the model learns to emit them into a real
# bid. Stripped from the text, counted so the caller can still see them.
_ANNOTATION = re.compile(
    r"\[[^\]]{0,160}?(VERIFY|CLIENT-SPECIFIC|CONSULTANT|TBD|PLACEHOLDER)[^\]]{0,160}?\]",
    re.I)


def count_annotations(text: str) -> int:
    """Unresolved drafting annotations in a passage. A document with any is not
    submission-ready, whatever else is true of it."""
    return len(_ANNOTATION.findall(text))


def strip_annotations(text: str) -> str:
    return re.sub(r"\s{2,}", " ", _ANNOTATION.sub("", text)).strip()

# A line repeated this often across a document is a running header or footer,
# whatever it happens to say.
_REPEAT_THRESHOLD = 3


def is_furniture(line: str) -> bool:
    text = line.strip()
    if len(text) < 3:
        return True
    return bool(
        _TOC_HEADER.match(text) or _TOC_LINE.search(text) or _PAGE_NO.match(text)
        or _PHONE.match(text) or _EMAIL_ONLY.match(text) or _URL_ONLY.match(text)
        or _ADDRESS.match(text))


def _drop_repeats(blocks: list[tuple[str, str]]) -> list[tuple[str, str]]:
    counts = Counter(text for _, text in blocks if 3 < len(text) < 120)
    repeated = {text for text, n in counts.items() if n >= _REPEAT_THRESHOLD}
    return [(kind, text) for kind, text in blocks
            if text not in repeated and not is_furniture(text)]


# ---------------------------------------------------------------------------
# Heading detection
# ---------------------------------------------------------------------------
_NUMBERED = re.compile(r"^\s*(\d{1,2}(\.\d{1,2})*|[IVXLC]+)[\.\)]?\s+\S")
_SECTION_WORDS = re.compile(
    r"^\s*(cover letter|transmittal|letter of transmittal|executive (summary|narrative)|"
    r"exective summary|company (background|profile|overview|history)|"
    r"qualifications?|firm qualifications|experience|"
    r"(project |implementation )?(approach|methodology|plan)|work plan|"
    r"project management|project governance|"
    r"staffing|key personnel|project team|resumes?|"
    r"references?|past performance|"
    r"technical (approach|requirements?|architecture|specifications?)|"
    r"(post.?implementation )?support|training|maintenance|"
    r"cost|pricing|price proposal|cost proposal|fee schedule|"
    r"risk|risk management|assumptions|exceptions|"
    r"terms and conditions|contract|appendix|attachment|exhibit)\b",
    re.I)


def _is_heading(text: str, style_name: str = "") -> bool:
    stripped = text.strip()
    if not stripped or len(stripped) > 110:
        return False
    if style_name and style_name.lower().startswith(("heading", "title", "subtitle")):
        return True
    if stripped.endswith((".", ",", ";")) and not _NUMBERED.match(stripped):
        return False
    words = stripped.split()
    if len(words) > 14:
        return False
    if _NUMBERED.match(stripped) and len(words) <= 12:
        return True
    if _SECTION_WORDS.match(stripped):
        return True
    letters = [c for c in stripped if c.isalpha()]
    if letters and sum(1 for c in letters if c.isupper()) / len(letters) > 0.85 \
            and len(words) <= 10:
        return True
    return False


# ---------------------------------------------------------------------------
# Extraction
# ---------------------------------------------------------------------------
def extract_docx(data: bytes) -> list[tuple[str, str]]:
    """Paragraphs and table cells in document order. Iterating the body element
    rather than paragraphs then tables keeps a table with the section it belongs
    to, instead of moving every table to the end of the document."""
    from docx.oxml.ns import qn
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    document = Document(io.BytesIO(data))
    blocks: list[tuple[str, str]] = []

    def emit(text: str, style: str) -> None:
        text = re.sub(r"\s+", " ", text).strip()
        if text:
            blocks.append(("H" if _is_heading(text, style) else "P", text))

    for child in document.element.body.iterchildren():
        if child.tag == qn("w:p"):
            paragraph = Paragraph(child, document)
            style = paragraph.style.name if paragraph.style is not None else ""
            emit(paragraph.text, style)
        elif child.tag == qn("w:tbl"):
            for row in Table(child, document).rows:
                # Merged cells repeat their text across the span; keep one copy.
                seen, ordered = set(), []
                for cell in row.cells:
                    value = cell.text.strip()
                    if value and value not in seen:
                        seen.add(value)
                        ordered.append(value)
                if ordered:
                    emit(" | ".join(ordered), "Table")
    return blocks


def extract_pdf(data: bytes) -> list[tuple[str, str]]:
    import pdfplumber

    blocks: list[tuple[str, str]] = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            for raw_line in (page.extract_text() or "").split("\n"):
                line = " ".join(raw_line.split())
                if line:
                    blocks.append(("H" if _is_heading(line) else "P", line))
    return blocks


def extract_xlsx(data: bytes) -> list[tuple[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    blocks: list[tuple[str, str]] = []
    try:
        for sheet in workbook.worksheets:
            blocks.append(("H", sheet.title))
            for row in sheet.iter_rows(values_only=True):
                cells = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if cells:
                    blocks.append(("P", " | ".join(cells)))
    finally:
        workbook.close()
    return blocks


def extract(filename: str, data: bytes) -> list[tuple[str, str]]:
    name = (filename or "").lower()
    try:
        if name.endswith(".pdf"):
            return extract_pdf(data)
        if name.endswith((".xlsx", ".xlsm", ".xls")):
            return extract_xlsx(data)
        if name.endswith(".docx"):
            return extract_docx(data)
    except ValidationFailed:
        raise
    except Exception as exc:
        log.warning("extraction failed for %s: %s", filename, exc)
        raise ValidationFailed(
            f"Could not read {filename}. The file may be corrupt or password protected.",
            {"reason": str(exc)},
        ) from exc
    if name.endswith(".doc"):
        # python-docx reads the OOXML .docx format only. Accepting .doc here
        # produced a corrupt-file error that blamed the document.
        raise ValidationFailed(
            f"{filename} is the legacy Word format. Save it as .docx and upload again.")
    raise ValidationFailed(
        f"Unsupported file type: {filename}. Upload DOCX, PDF, or XLSX.")


def plain_text(blocks: list[tuple[str, str]]) -> str:
    return "\n".join(text for _, text in blocks)


# ---------------------------------------------------------------------------
# Sectioning and chunking
# ---------------------------------------------------------------------------
def _sections(blocks: list[tuple[str, str]]) -> list[tuple[str, str]]:
    """Group blocks under the heading that introduced them."""
    sections: list[tuple[str, list[str]]] = []
    heading, body = "", []
    for kind, text in blocks:
        if kind == "H":
            if body:
                sections.append((heading, body))
            heading, body = text, []
        else:
            body.append(text)
    if body:
        sections.append((heading, body))
    return [(h, " ".join(b)) for h, b in sections if " ".join(b).strip()]


def _pack(sentences: list[str]) -> list[list[str]]:
    """Greedy pack into token-bounded groups, never splitting a sentence."""
    groups: list[list[str]] = []
    current: list[str] = []
    current_tokens = 0

    for sentence in sentences:
        tokens = est_tokens(sentence)
        if tokens > MAX_TOKENS:
            # A single sentence over the ceiling gets its own chunk rather than
            # being cut, because cutting it breaks rule 2.
            if current:
                groups.append(current)
                current, current_tokens = [], 0
            groups.append([sentence])
            continue
        if current and current_tokens + tokens > TARGET_TOKENS \
                and current_tokens >= MIN_TOKENS:
            groups.append(current)
            tail = current[-OVERLAP_SENTENCES:] if OVERLAP_SENTENCES else []
            current = list(tail)
            current_tokens = sum(est_tokens(x) for x in current)
        current.append(sentence)
        current_tokens += tokens

    if current:
        # Do not leave a runt: fold it back into the previous group.
        if groups and sum(est_tokens(x) for x in current) < MIN_TOKENS:
            groups[-1].extend(x for x in current if x not in groups[-1])
        else:
            groups.append(current)
    return groups


def chunk(blocks: list[tuple[str, str]]) -> list[dict]:
    """Section-aware, sentence-safe chunking.

    Each chunk records how its section tag was arrived at. 'body' means the text
    itself scored above threshold; 'smoothed' means it inherited the tag from
    confident neighbours. Retrieval ranks the first above the second, so this is
    not documentation.
    """
    blocks = _drop_repeats(blocks)
    pieces: list[dict] = []

    for heading, body in _sections(blocks):
        sentences = split_sentences(body)
        if not sentences:
            continue
        for group_index, group in enumerate(_pack(sentences)):
            raw = " ".join(group).strip()
            annotations = count_annotations(raw)
            text = tagger.normalize_dashes(strip_annotations(raw))
            if not text:
                continue
            # The heading rides with the first chunk so it reads standing alone,
            # rather than being repeated into every chunk of the section.
            scoped = f"{heading}. {text}" if (group_index == 0 and heading) else text

            section, score = tagger.tag_section(f"{heading} {text}" if heading else text)
            if section == DEFAULT_SECTION and heading:
                # A heading the scorer has no patterns for may still be one the
                # classifier recognises; this is how 'solution' and 'compliance'
                # are reached.
                from_heading = classifier.section_of(heading)
                if from_heading:
                    section, score = from_heading, tagger._MIN_SECTION_SCORE

            module, _ = tagger.tag_module(text)
            pieces.append({
                "section": section,
                "module": module,
                "text": scoped,
                "heading": heading,
                "tag_source": "body" if score >= tagger._MIN_SECTION_SCORE else "smoothed",
                "token_count": est_tokens(scoped),
                "annotations": annotations,
            })

    # Drop fragments that are both short and sentence-less. "18 Cost Proposal"
    # is a heading that survived furniture stripping, not answer content.
    pieces = [p for p in pieces
              if p["token_count"] >= MIN_TOKENS or len(p["text"].split()) >= 12]

    # Smoothing now runs inside real section boundaries: an unscored chunk
    # inherits from confident neighbours in the same document, rather than from
    # a table-of-contents line riding down the whole file.
    scored = [(p["section"] if p["tag_source"] == "body" else DEFAULT_SECTION, 0)
              for p in pieces]
    for piece, smoothed in zip(pieces, tagger.smooth_sections(scored)):
        if piece["tag_source"] != "body" and smoothed != DEFAULT_SECTION:
            piece["section"] = smoothed
    return pieces
