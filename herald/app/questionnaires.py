"""Excel questionnaire round-trip.

Import a vendor workbook, detect its structure, answer every question row from
the governed answer library and the retrieval index with a confidence score, and
write the answers back into the original file so its formatting, data validation,
and dropdowns survive intact.

Column detection combines header keywords with structural evidence: a column
carrying a list-type data validation is almost certainly the response column,
and that beats a keyword guess. Detected ranges are parsed properly from the
worksheet's validation squares rather than pattern-matched from a string.
"""
from __future__ import annotations

import io
import json
import logging

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from . import audit, classifier, documents, generation, opportunities
from .db import clob, cursor, transaction
from .errors import NotFound, ValidationFailed

log = logging.getLogger("harald.questionnaires")

import re

_QUESTION_HINT = re.compile(
    r"require|question|descriptio|feature|capabilit|functional|specif|criteri|item|need", re.I)
_RESPONSE_HINT = re.compile(
    r"response|rating|complian|\bmeets?\b|availab|vendor|\banswer\b|support|code|"
    r"\bmet\b|disposition|how (met|provided)", re.I)
_COMMENT_HINT = re.compile(
    r"comment|notes?|explanat|remark|approach|reference|detail|describe", re.I)

MAX_HEADER_SCAN_ROWS = 30
MAX_SCAN_COLUMNS = 60
MIN_QUESTION_LENGTH = 8
LOW_CONFIDENCE = 0.55

# How far below a candidate header we look for real question text, and how much
# of it we need before we believe the column is the question column.
MAX_BODY_PROBE_ROWS = 400
MIN_BODY_ROWS = 3

# An X-mark rating matrix: adjacent columns headed with short codes rather than
# prose. Nashua uses SUP / MOD / 3RD / CST / FUT / NS.
MIN_CODE_COLUMNS = 3
MAX_CODE_LABEL_LENGTH = 6

# Longest a cell can be and still be plausibly a column label rather than prose.
MAX_HEADER_LABEL_LENGTH = 60

# What a rating matrix expects in the chosen column. Nashua's instruction is
# "placing an X in the most appropriate column".
MATRIX_MARK = "X"

# How many rows may be in flight at once during a fill, and how often to log
# progress. A real requirements workbook runs to thousands of rows; answering
# them all at once takes the application down with it.
FILL_CONCURRENCY = 8
FILL_LOG_EVERY = 50


def _text(cell) -> str:
    if cell is None or cell.value is None:
        return ""
    return str(cell.value).strip()


def _list_validations(sheet: Worksheet) -> dict[int, list[str]]:
    """Map column index -> allowed values, from the sheet's own list validations.

    Ranges are parsed with openpyxl's range_boundaries, so a validation applied to
    C2:C500 correctly claims column C and nothing else. Formula-referenced lists
    (=Lists!$A$1:$A$6) are resolved by reading the referenced range.
    """
    columns: dict[int, list[str]] = {}
    workbook = sheet.parent

    for validation in sheet.data_validations.dataValidation:
        if validation.type != "list" or not validation.formula1:
            continue

        formula = validation.formula1.strip()
        values: list[str] = []
        if formula.startswith('"') and formula.endswith('"'):
            values = [v.strip() for v in formula[1:-1].split(",") if v.strip()]
        elif "!" in formula:
            try:
                sheet_ref, cell_ref = formula.lstrip("=").split("!", 1)
                source = workbook[sheet_ref.strip("'")]
                min_col, min_row, max_col, max_row = range_boundaries(
                    cell_ref.replace("$", "")
                )
                for row in source.iter_rows(min_row=min_row, max_row=max_row,
                                            min_col=min_col, max_col=max_col):
                    for cell in row:
                        if cell.value is not None and str(cell.value).strip():
                            values.append(str(cell.value).strip())
            except (KeyError, ValueError, TypeError) as exc:
                log.debug("could not resolve validation list %s: %s", formula, exc)
        if not values:
            continue

        for square in validation.sqref.ranges:
            bounds = range_boundaries(str(square))
            for column_index in range(bounds[0], bounds[2] + 1):
                columns.setdefault(column_index, values)

    return columns


def _looks_like_header(value: str) -> bool:
    """A column header is a label, not a sentence.

    Salem prints its rating legend above the real header, and one legend row
    reads "No: Feature/Function cannot be provided." That matches the question
    keywords on "Feature" and sits one row higher than the true header, so on
    keyword and body count alone it wins and the response column is lost. The
    thing that actually separates them is that legends are prose: they run long
    and they end in a full stop. Headers do neither.
    """
    if not value or len(value) > MAX_HEADER_LABEL_LENGTH:
        return False
    return not value.rstrip().endswith((".", "?", "!"))


def _body_rows(sheet: Worksheet, header_row: int, column_index: int,
               limit: int = MAX_BODY_PROBE_ROWS) -> int:
    """How many rows below ``header_row`` carry real question text in this column.

    This is the only signal that actually distinguishes a requirements column
    from a banner. Nashua's workbook opens with a merged title cell reading
    "Functional Requirements Matrix" in A2, which scores higher on keywords than
    the true header eight rows further down, and column A holds nothing below it.
    Counting the body is what tells the two apart.
    """
    last = min(sheet.max_row or 0, header_row + limit)
    found = 0
    for row_index in range(header_row + 1, last + 1):
        if len(_text(sheet.cell(row=row_index, column=column_index))) >= MIN_QUESTION_LENGTH:
            found += 1
    return found


def _code_columns(sheet: Worksheet, header_row: int,
                  scan_cols: int) -> dict[str, int]:
    """Detect an X-mark rating matrix: one column per response code.

    Salem and Jefferson use a single dropdown column. Nashua uses six adjacent
    columns headed SUP / MOD / 3RD / CST / FUT / NS, and the vendor marks an X
    under one of them. Both are ordinary layouts and neither is inferable from
    the other, so we look for a run of at least three adjacent columns whose
    headers are short codes rather than prose.
    """
    # The code strip does not have to sit on the header row we picked. Nashua's
    # question column starts at row 11 and its SUP/MOD/3RD/CST/FUT/NS strip is
    # on row 10, so we look above as well as below.
    for probe in (header_row, header_row - 1, header_row - 2, header_row - 3,
                  header_row + 1, header_row + 2):
        if probe < 1:
            continue
        run: list[tuple[int, str]] = []
        best: list[tuple[int, str]] = []
        for column_index in range(1, scan_cols + 1):
            value = _text(sheet.cell(row=probe, column=column_index))
            if value and len(value) <= MAX_CODE_LABEL_LENGTH and "\n" not in value:
                run.append((column_index, value))
                if len(run) > len(best):
                    best = list(run)
            else:
                run = []
        if len(best) >= MIN_CODE_COLUMNS:
            labels = [v for _, v in best]
            if len(set(labels)) == len(labels):
                return {v: c for c, v in best}
    return {}


def _detect(sheet: Worksheet) -> dict | None:
    """Find the header row and the question, response, and comment columns."""
    validated = _list_validations(sheet)
    scan_rows = min(sheet.max_row or 0, MAX_HEADER_SCAN_ROWS)
    scan_cols = min(sheet.max_column or 0, MAX_SCAN_COLUMNS)
    if not scan_rows or not scan_cols:
        return None

    # Score every plausible (header row, question column) pair rather than
    # picking a header row on keywords alone and then hoping a question column
    # falls out of it. The body count dominates deliberately: a column with 300
    # requirements under it is the question column whatever its header says, and
    # a keyword-perfect banner with an empty column under it is not.
    best = None
    best_score = -1
    for row_index in range(1, scan_rows + 1):
        headers: dict[int, str] = {}
        for column_index in range(1, scan_cols + 1):
            value = _text(sheet.cell(row=row_index, column=column_index))
            if value and len(value) <= 80:
                headers[column_index] = value
        if not headers:
            continue

        support = sum(
            1 for v in headers.values()
            if _RESPONSE_HINT.search(v) or _COMMENT_HINT.search(v)
        )
        for column_index in sorted(headers):
            if not _looks_like_header(headers[column_index]):
                continue
            # The body count is a qualifier, not the ranking term. Salem's
            # legend block sits above the real header and its definition column
            # holds six more rows of text than the requirements column below,
            # so ranking on body alone picks the legend. Keyword evidence ranks;
            # the body count only rejects banners and breaks ties.
            body = _body_rows(sheet, row_index, column_index)
            if body < MIN_BODY_ROWS:
                continue
            # Strict precedence: keyword match, then how much of the sheet the
            # candidate actually captures, then supporting headers. Letting
            # support outrank body lets a section divider partway down the sheet
            # beat the real header and silently truncate the import -- Nashua's
            # General Ledger tab lost 17 requirements exactly that way.
            score = body * 10 + support
            if _QUESTION_HINT.search(headers[column_index]):
                score += 100_000
            if score > best_score:
                best, best_score = (row_index, headers, column_index), score

    # Fall back to the old keyword-only rule for sheets whose question column is
    # genuinely sparse, so narrow workbooks that used to import still do.
    if best is None:
        for row_index in range(1, scan_rows + 1):
            headers = {}
            score = 0
            for column_index in range(1, scan_cols + 1):
                value = _text(sheet.cell(row=row_index, column=column_index))
                if not value or len(value) > 80:
                    continue
                headers[column_index] = value
                if _QUESTION_HINT.search(value):
                    score += 2
                if _RESPONSE_HINT.search(value) or _COMMENT_HINT.search(value):
                    score += 1
            question_col = next(
                (c for c, v in sorted(headers.items()) if _QUESTION_HINT.search(v)), None)
            if question_col is not None and score > best_score:
                best, best_score = (row_index, headers, question_col), score

    if best is None:
        return None

    header_row, headers, question_col = best

    code_columns = _code_columns(sheet, header_row, scan_cols)
    code_columns = {k: v for k, v in code_columns.items() if v != question_col}

    # Structural evidence wins: a validated column is the response column.
    response_col = next(
        (c for c in sorted(validated) if c != question_col and c in headers), None)
    if response_col is None:
        response_col = next(
            (c for c, v in sorted(headers.items())
             if c != question_col and _RESPONSE_HINT.search(v)), None)

    # A rating matrix has no single response column. Marking one would put the
    # answer in whichever rating column happened to sort first, which is a wrong
    # answer rather than a missing one.
    if code_columns and (response_col is None or response_col in code_columns.values()):
        response_col = None

    comment_col = next(
        (c for c, v in sorted(headers.items())
         if c not in (question_col, response_col)
         and c not in code_columns.values()
         and _COMMENT_HINT.search(v)), None)

    # Banner headings often sit a few rows above the row the requirements start
    # on. Nashua puts "ADDITIONAL COMMENTS" three rows up. Widen the search
    # rather than lose the comment column and with it every written answer.
    if comment_col is None:
        for probe in range(max(1, header_row - 4), header_row + 1):
            for column_index in range(1, scan_cols + 1):
                if column_index in (question_col, response_col):
                    continue
                if column_index in code_columns.values():
                    continue
                value = _text(sheet.cell(row=probe, column=column_index))
                if value and len(value) <= 80 and _COMMENT_HINT.search(value):
                    comment_col = column_index
                    break
            if comment_col is not None:
                break

    if code_columns:
        allowed = sorted(code_columns, key=lambda k: code_columns[k])
    else:
        allowed = validated.get(response_col, []) if response_col else []

    return {
        "header_row": header_row,
        "question_col": question_col,
        "response_col": response_col,
        "comment_col": comment_col,
        "code_columns": {k: get_column_letter(v) for k, v in code_columns.items()},
        "layout": "matrix" if code_columns else "single",
        "allowed_codes": allowed,
        "headers": {get_column_letter(c): v for c, v in headers.items()},
    }


def import_workbook(opp_id: int, source_doc_id: int, actor: str | None = None) -> dict:
    blob, filename = documents.get_blob(source_doc_id)
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise ValidationFailed(
            f"{filename} is not an Excel workbook. Import supports .xlsx and .xlsm.")

    workbook = load_workbook(io.BytesIO(blob), data_only=True)
    sheet_map: list[dict] = []
    items: list[dict] = []

    for sheet in workbook.worksheets:
        detected = _detect(sheet)
        if not detected:
            sheet_map.append({"sheet": sheet.title, "detected": False})
            continue

        question_letter = get_column_letter(detected["question_col"])
        response_letter = (get_column_letter(detected["response_col"])
                           if detected["response_col"] else None)
        comment_letter = (get_column_letter(detected["comment_col"])
                          if detected["comment_col"] else None)

        sheet_map.append({
            "sheet": sheet.title, "detected": True,
            "header_row": detected["header_row"], "question_col": question_letter,
            "response_col": response_letter, "comment_col": comment_letter,
            "code_columns": detected["code_columns"], "layout": detected["layout"],
            "allowed_codes": detected["allowed_codes"], "headers": detected["headers"],
        })

        for row_index in range(detected["header_row"] + 1, (sheet.max_row or 0) + 1):
            question = _text(sheet.cell(row=row_index, column=detected["question_col"]))
            if len(question) < MIN_QUESTION_LENGTH:
                continue
            items.append({
                "sheet": sheet.title, "row": row_index, "qcol": question_letter,
                "rcol": response_letter, "ccol": comment_letter,
                "question": question, "allowed": detected["allowed_codes"],
            })

    if not items:
        raise ValidationFailed(
            f"No question rows were detected in {filename}. The workbook may use an "
            f"unusual layout; the question column could not be identified.",
            {"sheet_map": sheet_map},
        )

    with transaction() as conn:
        cur = conn.cursor()
        out = cur.var(int)
        cur.execute(
            """INSERT INTO harald_questionnaires
                 (opp_id, source_doc_id, filename, sheet_map, status, item_count)
               VALUES (:opp, :doc, :fn, :map, 'imported', :count)
               RETURNING q_id INTO :out""",
            {"opp": opp_id, "doc": source_doc_id, "fn": filename,
             "map": json.dumps(sheet_map), "count": len(items), "out": out},
        )
        q_id = out.getvalue()[0]
        # :row is Oracle reserved (ORA-01745); keep item binds on b_* names.
        cur.executemany(
            """INSERT INTO harald_questionnaire_items
                 (q_id, opp_id, sheet_name, row_index, question_col, response_col,
                  comment_col, question_text, allowed_codes, sort_order)
               VALUES (:b_q, :b_opp, :b_sheet, :b_row, :b_qcol, :b_rcol, :b_ccol,
                       :b_text, :b_codes, :b_ord)""",
            [
                {
                    "b_q": q_id,
                    "b_opp": opp_id,
                    "b_sheet": item["sheet"],
                    "b_row": item["row"],
                    "b_qcol": item["qcol"],
                    "b_rcol": item["rcol"],
                    "b_ccol": item["ccol"],
                    "b_text": item["question"],
                    "b_codes": json.dumps(item["allowed"]),
                    "b_ord": i,
                }
                for i, item in enumerate(items)
            ],
        )

    audit.record(actor, "questionnaire.import", "questionnaire", q_id,
                 {"file": filename, "items": len(items)})
    log.info("imported questionnaire q_id=%s file=%s items=%s", q_id, filename, len(items))
    return {"q_id": q_id, "filename": filename, "item_count": len(items),
            "sheet_map": sheet_map}


def list_for_opportunity(opp_id: int) -> list[dict]:
    with cursor() as cur:
        cur.execute(
            """SELECT q.q_id, q.filename, q.status, q.item_count, q.fill_error, q.created_at,
                      (SELECT COUNT(*) FROM harald_questionnaire_items i
                        WHERE i.q_id = q.q_id AND i.status <> 'todo'),
                      (SELECT COUNT(*) FROM harald_questionnaire_items i
                        WHERE i.q_id = q.q_id AND i.status = 'needs_review'),
                      q.source_doc_id
               FROM harald_questionnaires q WHERE q.opp_id = :o
               ORDER BY q.created_at DESC""",
            {"o": opp_id},
        )
        return [
            {"q_id": r[0], "filename": r[1], "status": r[2], "item_count": r[3],
             "fill_error": r[4], "created_at": r[5].isoformat() if r[5] else None,
             "answered": r[6], "needs_review": r[7], "source_doc_id": r[8]}
            for r in cur.fetchall()
        ]


def get(q_id: int) -> dict:
    with cursor() as cur:
        cur.execute(
            """SELECT q_id, opp_id, source_doc_id, filename, sheet_map, status,
                      item_count, fill_error
               FROM harald_questionnaires WHERE q_id = :q""",
            {"q": q_id},
        )
        row = cur.fetchone()
        if not row:
            raise NotFound(f"Questionnaire {q_id} not found.")
        try:
            sheet_map = json.loads(clob(row[4])) if row[4] else []
        except (json.JSONDecodeError, TypeError):
            sheet_map = []
        result = {"q_id": row[0], "opp_id": row[1], "source_doc_id": row[2],
                  "filename": row[3], "sheet_map": sheet_map, "status": row[5],
                  "item_count": row[6], "fill_error": row[7], "items": []}

        cur.execute(
            """SELECT qi_id, sheet_name, row_index, question_text, allowed_codes,
                      response_code, response_text, confidence, status, owner,
                      source_answer_id, response_col, comment_col
               FROM harald_questionnaire_items WHERE q_id = :q ORDER BY sort_order""",
            {"q": q_id},
        )
        for r in cur.fetchall():
            try:
                allowed = json.loads(clob(r[4])) if r[4] else []
            except (json.JSONDecodeError, TypeError):
                allowed = []
            result["items"].append({
                "qi_id": r[0], "sheet": r[1], "row": r[2], "question": clob(r[3]),
                "allowed_codes": allowed, "response_code": r[5],
                "response_text": clob(r[6]), "confidence": r[7], "status": r[8],
                "owner": r[9], "source_answer_id": r[10],
                # Carried here so export() does not re-query per row: a 600-row
                # agency workbook was 600 sequential pool acquisitions.
                "response_col": r[11], "comment_col": r[12],
            })
    return result


def set_status(q_id: int, status: str, error: str | None = None) -> None:
    with transaction() as conn:
        conn.cursor().execute(
            "UPDATE harald_questionnaires SET status = :s, fill_error = :e WHERE q_id = :q",
            {"s": status, "e": (error or "")[:2000] or None, "q": q_id},
        )


def update_item(qi_id: int, payload: dict, actor: str | None = None) -> None:
    allowed = {"response_code", "response_text", "status", "owner"}
    columns = {k: v for k, v in payload.items() if k in allowed}
    if "status" in columns and columns["status"] not in (
        "todo", "drafted", "needs_review", "approved"
    ):
        raise ValidationFailed("Invalid questionnaire item status.")
    if not columns:
        return
    with transaction() as conn:
        cur = conn.cursor()
        assignments = ", ".join(f"{col} = :{col}" for col in columns)
        cur.execute(
            f"UPDATE harald_questionnaire_items SET {assignments} WHERE qi_id = :qi",
            {**columns, "qi": qi_id},
        )
        if cur.rowcount == 0:
            raise NotFound(f"Questionnaire item {qi_id} not found.")


def _persist_fill(qi_id: int, result: dict) -> None:
    status = "needs_review" if result["confidence"] < LOW_CONFIDENCE else "drafted"
    with transaction() as conn:
        conn.cursor().execute(
            """UPDATE harald_questionnaire_items
               SET response_code = :code, response_text = :text, confidence = :conf,
                   source_answer_id = :src, status = :status
               WHERE qi_id = :qi AND status <> 'approved'""",
            {"code": result["response_code"], "text": result["response_text"],
             "conf": result["confidence"], "src": result.get("source_answer_id"),
             "status": status, "qi": qi_id},
        )


async def fill(q_id: int, actor: str | None = None, *, redo: bool = False) -> None:
    """Answer every unapproved row, a bounded number at a time.

    The comment this docstring replaced said concurrency was bounded inside the
    model client. It is not. A bare gather over every pending row opened one
    request per row, which was survivable at the 45 rows a mis-parsed workbook
    produced and is not survivable at the 3,041 a correct parse produces: the
    fan-out saturated the event loop and the whole application stopped
    answering, including the pages an operator would use to see why.

    A semaphore caps in-flight work. Progress is logged so a long fill is
    observable rather than indistinguishable from a hang.
    """
    import asyncio

    # Any render held from before this fill is now wrong.
    _EXPORT_CACHE.pop(q_id, None)

    try:
        set_status(q_id, "filling")
        questionnaire = get(q_id)
        # Resume rather than restart. A fill over a real requirements workbook
        # runs for over an hour, so a container restart part-way through used to
        # mean redoing every row that already had an answer -- the second run
        # spent an hour reproducing work before reaching the rows that were
        # actually missing. Skip rows that already carry an answer unless the
        # caller asks for a rewrite.
        pending = [
            i for i in questionnaire["items"]
            if i["status"] != "approved"
            and (redo or not (i.get("response_code") or i.get("response_text")))
        ]
        if not pending:
            set_status(q_id, "filled")
            log.info("fill q_id=%s nothing to do, every row already answered", q_id)
            return
        rfp_text, parsed_fields = "", {}
        if questionnaire.get("opp_id"):
            try:
                rfp_text, parsed_fields = opportunities.grounding_context(
                    opportunities.get(questionnaire["opp_id"]),
                )
            except Exception:
                log.debug("fill: no opp grounding for q_id=%s", q_id)

        gate = asyncio.Semaphore(FILL_CONCURRENCY)
        done = 0
        total = len(pending)

        async def answer_one(item: dict) -> None:
            nonlocal done
            async with gate:
                await _answer_row(item, rfp_text, parsed_fields)
            done += 1
            if done % FILL_LOG_EVERY == 0 or done == total:
                log.info("fill q_id=%s %s/%s rows", q_id, done, total)

        async def _answer_row(item: dict, rfp_text: str, parsed_fields: dict) -> None:
            module = classifier.module_of(item["question"])
            try:
                result = await generation.answer_question(
                    item["question"], module, item["allowed_codes"] or None,
                    rfp_text=rfp_text, parsed_fields=parsed_fields,
                )
            except Exception as exc:
                log.warning("fill failed qi_id=%s: %s", item["qi_id"], exc)
                # Never leave a blank "can't do it" cell — flag for a human.
                codes = item.get("allowed_codes") or []
                result = {
                    "response_code": generation._preferred_constructive_code(codes)
                    if codes else "",
                    "response_text": (
                        "[NEEDS HUMAN: fill failed automatically — "
                        f"{type(exc).__name__}. Complete this row manually.]"
                    ),
                    "confidence": 0.0,
                    "source_answer_id": None,
                }
            _persist_fill(item["qi_id"], result)

        log.info("fill q_id=%s starting %s rows, %s at a time",
                 q_id, total, FILL_CONCURRENCY)
        await asyncio.gather(*(answer_one(item) for item in pending))
        set_status(q_id, "filled")
        audit.record(actor, "questionnaire.fill", "questionnaire", q_id,
                     {"items": len(pending)})
    except Exception as exc:
        log.exception("questionnaire fill failed q_id=%s", q_id)
        set_status(q_id, "error", str(exc))


# What each rating means, independent of the label an agency prints on it.
# Left column is the concept; the lists are the spellings seen in real
# solicitations. Salem and Jefferson use words, Nashua uses three-letter codes.
_CODE_SENSE = {
    "delivered":     ["sup", "standard", "s", "yes", "y", "out of the box", "ootb"],
    "configured":    ["mod", "configuration", "configurable", "c", "config"],
    "third_party":   ["3rd", "third party", "thirdparty", "t", "partner"],
    "custom":        ["cst", "customization", "customisation", "modification", "custom"],
    "future":        ["fut", "future release", "future", "f", "roadmap"],
    "not_supported": ["ns", "not available", "not supported", "no", "n", "no bid"],
}


def _sense_of(code: str) -> str | None:
    probe = re.sub(r"[^a-z0-9 ]+", "", (code or "").strip().lower())
    for sense, spellings in _CODE_SENSE.items():
        if probe in spellings:
            return sense
    return None


def _translate_code(code: str, code_columns: dict) -> str | None:
    """Re-spell a rating in the vocabulary this workbook actually uses.

    A rating carries meaning, not just a label. Dropping an answer because the
    agency spells "supported out of the box" as SUP rather than Standard turns
    a scored row into a blank one, and a blank scores zero.
    """
    sense = _sense_of(code)
    if not sense:
        return None
    for candidate in code_columns:
        if _sense_of(candidate) == sense:
            return candidate
    return None


def _write_cell(sheet: Worksheet, coordinate: str, value) -> bool:
    """Write ``value`` into ``coordinate``, resolving merged ranges.

    openpyxl exposes every cell of a merged range except the top-left anchor as
    a read-only ``MergedCell``; assigning to one raises AttributeError. Agency
    requirements workbooks merge response and comment columns constantly, so a
    blind write kills the whole export over a single cell. We redirect the
    write to the range's anchor, which is the cell the merge actually displays,
    and return False for the cases we cannot place so the caller can log them
    rather than lose the file.
    """
    try:
        cell = sheet[coordinate]
    except (ValueError, IndexError):
        return False

    if isinstance(cell, MergedCell):
        anchor = None
        for merged in sheet.merged_cells.ranges:
            if coordinate in merged:
                anchor = sheet.cell(row=merged.min_row, column=merged.min_col)
                break
        if anchor is None:
            return False
        cell = anchor

    try:
        cell.value = value
    except AttributeError:
        return False
    return True


# Rendering an answered workbook is an openpyxl load-and-save of the agency's
# own file. Nashua's Appendix A is 9.4 MB and its packet holds four of them, so
# a materials.zip download took 176 seconds and looked to the person waiting
# like a hang. The render is pure: the same questionnaire at the same answer
# count produces the same bytes, so it only has to happen once per container.
EXPORT_CACHE_MAX = 12
_EXPORT_CACHE: dict[int, tuple[tuple, bytes, str]] = {}


def _export_version(q_id: int) -> tuple:
    """Cheap fingerprint of everything the render reads."""
    with cursor() as cur:
        cur.execute(
            """SELECT q.status, q.item_count, q.source_doc_id,
                      (SELECT COUNT(*) FROM harald_questionnaire_items i
                        WHERE i.q_id = q.q_id AND i.status <> 'todo'),
                      (SELECT NVL(MAX(i.qi_id), 0) FROM harald_questionnaire_items i
                        WHERE i.q_id = q.q_id),
                      (SELECT NVL(SUM(ORA_HASH(NVL(i.response_code, '-'))), 0)
                         FROM harald_questionnaire_items i WHERE i.q_id = q.q_id)
               FROM harald_questionnaires q WHERE q.q_id = :q""",
            {"q": q_id},
        )
        row = cur.fetchone()
    if not row:
        raise NotFound(f"Questionnaire {q_id} not found.")
    return tuple(row)


def export(q_id: int) -> tuple[bytes, str]:
    """Render the answered workbook, reusing the last render when nothing moved."""
    try:
        version = _export_version(q_id)
    except NotFound:
        raise
    except Exception:
        log.exception("questionnaire q_id=%s version probe failed; rendering", q_id)
        return _render_export(q_id)

    cached = _EXPORT_CACHE.get(q_id)
    if cached and cached[0] == version:
        log.info("questionnaire q_id=%s export served from cache", q_id)
        return cached[1], cached[2]

    blob, filename = _render_export(q_id)
    _EXPORT_CACHE[q_id] = (version, blob, filename)
    while len(_EXPORT_CACHE) > EXPORT_CACHE_MAX:
        _EXPORT_CACHE.pop(next(iter(_EXPORT_CACHE)))
    return blob, filename


def _render_export(q_id: int) -> tuple[bytes, str]:
    """Write answers back into the original workbook. Loading without data_only
    preserves formulas, styles, and data validations, so the exported file is the
    agency's own workbook with iteria's answers in it."""
    questionnaire = get(q_id)
    blob, filename = documents.get_blob(questionnaire["source_doc_id"])
    workbook = load_workbook(io.BytesIO(blob))

    # Per-sheet layout, so a rating matrix can be written the way its own
    # workbook expects rather than as text in a column that does not exist.
    #
    # The stored sheet_map is a hint, not the source of truth. Nashua's came
    # back empty and every rating mark was silently dropped: 3,130 comments
    # written into column J and not one X in D through I, which is a workbook
    # that scores zero while looking filled. The workbook is already open here,
    # so re-detecting costs nothing and cannot go stale.
    layouts = {
        entry.get("sheet"): entry
        for entry in (questionnaire.get("sheet_map") or [])
        if isinstance(entry, dict)
    }
    if any(not (layouts.get(s.title) or {}).get("code_columns")
           for s in workbook.worksheets):
        # Detection must see the same workbook the import saw. import_workbook
        # loads with data_only=True so formula cells yield their cached value;
        # export loads without it so formulas survive the round trip. Detecting
        # on the writable copy finds nothing, because the question column reads
        # back as "=IF(Systems!B5=..." instead of the requirement text. Two
        # loads, one to look and one to write.
        probe = load_workbook(io.BytesIO(blob), data_only=True)
        for sheet in probe.worksheets:
            entry = layouts.get(sheet.title) or {}
            if entry.get("code_columns"):
                continue
            detected = _detect(sheet)
            if detected and detected.get("code_columns"):
                entry = dict(entry)
                entry["code_columns"] = detected["code_columns"]
                entry.setdefault("allowed_codes", detected["allowed_codes"])
                layouts[sheet.title] = entry
        probe.close()
        recovered = sum(1 for v in layouts.values() if v.get("code_columns"))
        if recovered:
            log.info("export q_id=%s recovered rating columns on %s sheet(s)",
                     q_id, recovered)

    written = 0
    skipped: list[str] = []
    for item in questionnaire["items"]:
        if item["sheet"] not in workbook.sheetnames:
            continue
        sheet = workbook[item["sheet"]]
        response_col = item.get("response_col")
        comment_col = item.get("comment_col")
        code_columns = (layouts.get(item["sheet"]) or {}).get("code_columns") or {}

        code = item["response_code"]
        if code_columns and code and code not in code_columns:
            # The row carries a code from a different workbook's legend, because
            # allowed_codes did not reach the model. Salem says "Standard";
            # Nashua's own legend says "SUP". Same meaning, different agency.
            # Translate rather than drop the answer on the floor.
            mapped = _translate_code(code, code_columns)
            if mapped:
                item = {**item, "response_code": mapped}
            else:
                skipped.append(
                    f"{item['sheet']}!row{item['row']} (no column for {code!r})")
                item = {**item, "response_code": ""}

        if code_columns and item["response_code"]:
            # An X-mark matrix. The RFP treats more than one mark on a row as a
            # non-response, so clear the whole strip before marking, otherwise a
            # re-export after a changed answer leaves two X's and scores zero.
            placed = False
            for code, column in code_columns.items():
                target = f"{column}{item['row']}"
                if code == item["response_code"]:
                    placed = _write_cell(sheet, target, MATRIX_MARK)
                    if not placed:
                        skipped.append(f"{item['sheet']}!{target}")
                else:
                    _write_cell(sheet, target, None)
            if placed:
                written += 1
            elif item["response_code"] not in code_columns:
                skipped.append(
                    f"{item['sheet']}!row{item['row']} (no column for "
                    f"{item['response_code']!r})")
        elif response_col and item["response_code"]:
            target = f"{response_col}{item['row']}"
            if _write_cell(sheet, target, item["response_code"]):
                written += 1
            else:
                skipped.append(f"{item['sheet']}!{target}")

        if comment_col and item["response_text"]:
            target = f"{comment_col}{item['row']}"
            if not _write_cell(sheet, target, item["response_text"]):
                skipped.append(f"{item['sheet']}!{target}")

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # A workbook with no mapped response column -- a cost worksheet is the usual
    # case -- writes nothing, and the file handed back is the agency's original.
    # Calling that "exported" is the most expensive lie the system can tell,
    # because it reads as done on the screen the operator checks before
    # submitting. Report what actually happened instead. The answers are kept:
    # the import path is deliberately left intact so this material is already
    # structured on the day there is enough pricing data to answer it properly.
    if written == 0:
        set_status(q_id, "filled",
                   "No response column is mapped on this workbook, so none of "
                   f"the {len(questionnaire['items'])} answers were written. "
                   "The downloaded file is the agency's original. Cost and "
                   "pricing worksheets normally land here and need a human.")
        log.warning("questionnaire q_id=%s exported 0 cells (file=%s)",
                    q_id, filename)
    else:
        set_status(q_id, "exported",
                   f"{len(skipped)} cell(s) could not be placed." if skipped else None)

    stem = filename.rsplit(".", 1)[0]
    if skipped:
        log.warning("questionnaire q_id=%s could not place %s cell(s): %s",
                    q_id, len(skipped), ", ".join(skipped[:20]))
    log.info("exported questionnaire q_id=%s cells_written=%s skipped=%s",
             q_id, written, len(skipped))
    return buffer.read(), f"{stem}_iteria_response.xlsx"
