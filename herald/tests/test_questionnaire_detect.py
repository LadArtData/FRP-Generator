"""Regression tests for questionnaire structure detection.

Nashua's Appendix A workbook imported 45 requirements out of 23 functional tabs
holding roughly 3,000. The narrative then attested that the worksheets had been
"completed by iteria functional leads". Two independent detector faults caused
it, and both are pinned here.
"""

import io

import pytest
from openpyxl import Workbook, load_workbook

from app.questionnaires import (
    MATRIX_MARK,
    MIN_QUESTION_LENGTH,
    _code_columns,
    _detect,
    _looks_like_header,
    _text,
)


def _count_items(sheet, detected):
    return sum(
        1 for row in range(detected["header_row"] + 1, (sheet.max_row or 0) + 1)
        if len(_text(sheet.cell(row=row, column=detected["question_col"]))) >= MIN_QUESTION_LENGTH
    )


def _nashua_style_sheet(rows=60):
    """A banner title, a legend block, an X-mark rating strip, then the body."""
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "CITY OF NASHUA"
    sheet["A2"] = "Functional Requirements Matrix"
    sheet["D2"] = "Complete the worksheet by placing an X in the column"
    sheet["A4"] = "Vendor Name:"
    sheet["A8"] = "GENERAL LEDGER MANAGEMENT"
    sheet["D8"] = "RATING RESPONSE"
    sheet["J8"] = "ADDITIONAL COMMENTS"
    for column, label in zip("DEFGHI", ["SUP", "MOD", "3RD", "CST", "FUT", "NS"]):
        sheet[f"{column}10"] = label
    sheet["C11"] = "System Setup Requirements Section"
    for offset in range(rows):
        sheet[f"B{12 + offset}"] = f"GL.{offset + 1}"
        sheet[f"C{12 + offset}"] = f"The system shall support requirement {offset + 1}"
    return workbook, sheet


def _salem_style_sheet(rows=40):
    """A rating legend in prose above the real header row."""
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Indicator"
    sheet["B1"] = "Definition"
    sheet["C1"] = "Instruction"
    legend = [
        ("S", "Standard: Feature/Function is included in the base product."),
        ("F", "Future: Feature/Function will be available in a future release."),
        ("C", "Customization: Feature/Function is not included."),
        ("T", "Third Party: Feature/Function is not included."),
        ("N", "No: Feature/Function cannot be provided."),
    ]
    for index, (code, definition) in enumerate(legend, start=2):
        sheet[f"A{index}"] = code
        sheet[f"B{index}"] = definition
    sheet["A7"] = "General Ledger and Financial Reporting"
    sheet["A8"] = "Req #"
    sheet["B8"] = "Description of Requirement"
    sheet["C8"] = "Criticality"
    sheet["D8"] = "Vendor Response"
    sheet["E8"] = "Comments"
    for offset in range(rows):
        sheet[f"A{9 + offset}"] = f"GL.{offset + 1}"
        sheet[f"B{9 + offset}"] = f"The system shall support requirement {offset + 1}"
    return workbook, sheet


# ---------------------------------------------------------------------------
# Fault 1: a banner title outranked the real header, and the body was lost.
# ---------------------------------------------------------------------------

def test_banner_title_does_not_win_over_the_real_header():
    workbook, sheet = _nashua_style_sheet(rows=60)
    detected = _detect(sheet)
    assert detected is not None
    # A2 "Functional Requirements Matrix" is keyword-perfect and has nothing
    # under it. C11 is the column with 60 requirements below it.
    assert detected["header_row"] == 11
    assert detected["question_col"] == 3


def test_banner_title_regression_full_body_is_extracted():
    workbook, sheet = _nashua_style_sheet(rows=300)
    detected = _detect(sheet)
    assert _count_items(sheet, detected) == 300, "the whole sheet must import"


# ---------------------------------------------------------------------------
# Fault 2: an X-mark rating matrix has no single response column.
# ---------------------------------------------------------------------------

def test_rating_strip_is_detected_as_code_columns():
    workbook, sheet = _nashua_style_sheet()
    detected = _detect(sheet)
    assert detected["layout"] == "matrix"
    assert detected["code_columns"] == {
        "SUP": "D", "MOD": "E", "3RD": "F", "CST": "G", "FUT": "H", "NS": "I",
    }
    assert detected["allowed_codes"] == ["SUP", "MOD", "3RD", "CST", "FUT", "NS"]


def test_matrix_sheet_exposes_no_single_response_column():
    """Picking one would put every answer in whichever rating sorted first."""
    workbook, sheet = _nashua_style_sheet()
    detected = _detect(sheet)
    assert detected["response_col"] is None


def test_matrix_sheet_still_finds_the_comment_column_above_the_header():
    workbook, sheet = _nashua_style_sheet()
    detected = _detect(sheet)
    assert detected["comment_col"] == 10  # column J, three rows above


def test_code_columns_found_when_strip_sits_above_the_header_row():
    workbook, sheet = _nashua_style_sheet()
    assert _code_columns(sheet, 11, 20) == {
        "SUP": 4, "MOD": 5, "3RD": 6, "CST": 7, "FUT": 8, "NS": 9,
    }


# ---------------------------------------------------------------------------
# The prose-legend fault that a naive body-count fix reintroduces.
# ---------------------------------------------------------------------------

def test_prose_legend_row_does_not_win_over_the_real_header():
    """"No: Feature/Function cannot be provided." matches the question keywords
    and sits above the real header with a higher body count. It is a sentence,
    and headers are not sentences."""
    workbook, sheet = _salem_style_sheet()
    detected = _detect(sheet)
    assert detected["header_row"] == 8
    assert detected["question_col"] == 2


def test_single_column_layout_keeps_response_and_comment_columns():
    workbook, sheet = _salem_style_sheet()
    detected = _detect(sheet)
    assert detected["layout"] == "single"
    assert detected["response_col"] == 4   # D, Vendor Response
    assert detected["comment_col"] == 5    # E, Comments
    assert detected["code_columns"] == {}


@pytest.mark.parametrize("value,expected", [
    ("Description of Requirement", True),
    ("Req #", True),
    ("System Setup Requirements Section", True),
    ("No: Feature/Function cannot be provided.", False),
    ("Standard: Feature/Function is included in the base product.", False),
    ("Is this supported?", False),
    ("", False),
    ("x" * 80, False),
])
def test_looks_like_header(value, expected):
    assert _looks_like_header(value) is expected


# ---------------------------------------------------------------------------
# Export: the mark lands in the right column and nowhere else.
# ---------------------------------------------------------------------------

def _place(sheet, code_columns, row, chosen):
    from app.questionnaires import _write_cell
    for code, column in code_columns.items():
        target = f"{column}{row}"
        _write_cell(sheet, target, MATRIX_MARK if code == chosen else None)


def test_matrix_export_marks_one_column_only():
    workbook, sheet = _nashua_style_sheet()
    columns = {"SUP": "D", "MOD": "E", "3RD": "F", "CST": "G", "FUT": "H", "NS": "I"}
    _place(sheet, columns, 12, "MOD")
    assert sheet["E12"].value == MATRIX_MARK
    assert [sheet[f"{c}12"].value for c in "DFGHI"] == [None] * 5


def test_matrix_export_clears_a_previous_mark_on_reexport():
    """More than one mark on a row is scored as a non-response."""
    workbook, sheet = _nashua_style_sheet()
    columns = {"SUP": "D", "MOD": "E", "3RD": "F", "CST": "G", "FUT": "H", "NS": "I"}
    _place(sheet, columns, 12, "SUP")
    _place(sheet, columns, 12, "CST")
    marked = [c for c in "DEFGHI" if sheet[f"{c}12"].value == MATRIX_MARK]
    assert marked == ["G"]


def test_matrix_export_survives_round_trip():
    workbook, sheet = _nashua_style_sheet()
    columns = {"SUP": "D", "MOD": "E", "3RD": "F", "CST": "G", "FUT": "H", "NS": "I"}
    _place(sheet, columns, 12, "SUP")
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    assert load_workbook(buffer).active["D12"].value == MATRIX_MARK


# ---------------------------------------------------------------------------
# Fill concurrency. A bare gather over 3,041 rows took the application down.
# ---------------------------------------------------------------------------

def test_fill_bounds_concurrency():
    import inspect
    from app import questionnaires
    src = inspect.getsource(questionnaires.fill)
    assert "asyncio.Semaphore(FILL_CONCURRENCY)" in src, (
        "fill must cap in-flight rows; an unbounded gather over a real "
        "requirements workbook saturates the event loop"
    )
    gate = src.index("async with gate")
    call = src.index("_answer_row(item")
    assert gate < call, "the semaphore must be held across the model call"


def test_fill_concurrency_is_sane():
    from app import questionnaires
    assert 1 <= questionnaires.FILL_CONCURRENCY <= 32


def test_fill_logs_progress():
    import inspect
    from app import questionnaires
    src = inspect.getsource(questionnaires.fill)
    assert "FILL_LOG_EVERY" in src, "a long fill must be distinguishable from a hang"


# ---------------------------------------------------------------------------
# Resume. A container restart mid-fill used to mean redoing every answered row.
# ---------------------------------------------------------------------------

def test_fill_resumes_instead_of_restarting():
    import inspect
    from app import questionnaires
    src = inspect.getsource(questionnaires.fill)
    assert "redo or not (i.get(\"response_code\") or i.get(\"response_text\"))" in src, (
        "a resumed fill must skip rows that already carry an answer"
    )


def test_fill_accepts_a_redo_flag():
    import inspect
    from app import questionnaires
    sig = inspect.signature(questionnaires.fill)
    assert "redo" in sig.parameters
    assert sig.parameters["redo"].default is False


def test_fill_route_exposes_redo():
    import inspect
    from app import main
    src = inspect.getsource(main.fill_questionnaire)
    assert "redo" in src


# ---------------------------------------------------------------------------
# Export wrote 3,130 comments and zero rating marks: a workbook that scores
# zero while looking filled. Two causes, both pinned here.
# ---------------------------------------------------------------------------

def test_detection_survives_a_workbook_loaded_for_writing():
    """import loads data_only=True, export does not. Detecting on the writable
    copy sees formulas instead of requirement text and finds nothing."""
    import io as _io
    from openpyxl import load_workbook as _load
    workbook, sheet = _nashua_style_sheet(rows=40)
    sheet["C20"] = "=IF(Systems!B20=\"\",\"\",Systems!B20)"
    buffer = _io.BytesIO()
    workbook.save(buffer)

    buffer.seek(0)
    probe = _load(buffer, data_only=True)
    assert _detect(probe.active)["code_columns"], "detection copy must find the strip"


def test_translate_code_maps_between_agency_legends():
    from app.questionnaires import _translate_code
    nashua = {"SUP": "D", "MOD": "E", "3RD": "F", "CST": "G", "FUT": "H", "NS": "I"}
    assert _translate_code("Standard", nashua) == "SUP"
    assert _translate_code("Configuration", nashua) == "MOD"
    assert _translate_code("Third Party", nashua) == "3RD"
    assert _translate_code("Modification", nashua) == "CST"
    assert _translate_code("Future Release", nashua) == "FUT"
    assert _translate_code("Not Available", nashua) == "NS"
    assert _translate_code("No Bid", nashua) == "NS"


def test_translate_code_refuses_a_meaning_it_does_not_know():
    from app.questionnaires import _translate_code
    assert _translate_code("banana", {"SUP": "D", "NS": "I"}) is None


def test_translate_code_is_identity_within_one_legend():
    from app.questionnaires import _translate_code
    salem = {"Standard": "D", "Configuration": "E", "Not Available": "F"}
    assert _translate_code("Standard", salem) == "Standard"


def test_export_recovers_layout_when_sheet_map_is_empty():
    """The stored sheet_map came back empty and every mark was dropped. The
    export must not depend on it."""
    import inspect
    from app import questionnaires
    src = inspect.getsource(questionnaires._render_export)
    assert "data_only=True" in src, "export needs a detection copy"
    assert "_detect(sheet)" in src, "export must re-detect when sheet_map is thin"


# ---------------------------------------------------------------------------
# Export caching. A materials.zip download took 176 seconds because every
# questionnaire in the packet re-ran an openpyxl load-and-save of the agency's
# own workbook, one of which is 9.4 MB.
# ---------------------------------------------------------------------------

@pytest.fixture
def cache(monkeypatch):
    from app import questionnaires
    questionnaires._EXPORT_CACHE.clear()
    calls = {"render": 0}

    def fake_render(q_id):
        calls["render"] += 1
        return b"xlsx-bytes-%d" % q_id, f"q{q_id}_iteria_response.xlsx"

    monkeypatch.setattr(questionnaires, "_render_export", fake_render)
    yield questionnaires, calls
    questionnaires._EXPORT_CACHE.clear()


def test_second_export_of_an_unchanged_questionnaire_is_not_rendered(cache, monkeypatch):
    questionnaires, calls = cache
    monkeypatch.setattr(questionnaires, "_export_version", lambda q: ("filled", 3041, 5260, 3041, 99, 7))
    first = questionnaires.export(41)
    second = questionnaires.export(41)
    assert first == second
    assert calls["render"] == 1


def test_a_refilled_questionnaire_is_rendered_again(cache, monkeypatch):
    questionnaires, calls = cache
    version = {"v": ("filled", 3041, 5260, 3041, 99, 7)}
    monkeypatch.setattr(questionnaires, "_export_version", lambda q: version["v"])
    questionnaires.export(41)
    version["v"] = ("filled", 3041, 5260, 3041, 99, 8)   # a response code changed
    questionnaires.export(41)
    assert calls["render"] == 2


def test_answer_count_change_invalidates_the_render(cache, monkeypatch):
    questionnaires, calls = cache
    version = {"v": ("filling", 3041, 5260, 12, 99, 7)}
    monkeypatch.setattr(questionnaires, "_export_version", lambda q: version["v"])
    questionnaires.export(41)
    version["v"] = ("filled", 3041, 5260, 3041, 99, 7)
    questionnaires.export(41)
    assert calls["render"] == 2


def test_cache_does_not_grow_without_bound(cache, monkeypatch):
    questionnaires, calls = cache
    monkeypatch.setattr(questionnaires, "_export_version", lambda q: ("filled", q))
    for q_id in range(questionnaires.EXPORT_CACHE_MAX + 5):
        questionnaires.export(q_id)
    assert len(questionnaires._EXPORT_CACHE) <= questionnaires.EXPORT_CACHE_MAX


def test_two_questionnaires_do_not_share_a_render(cache, monkeypatch):
    questionnaires, calls = cache
    monkeypatch.setattr(questionnaires, "_export_version", lambda q: ("filled", q))
    a, _ = questionnaires.export(41)
    b, _ = questionnaires.export(21)
    assert a != b


def test_a_broken_version_probe_still_serves_the_file(cache, monkeypatch):
    """A cache is an optimisation. It must never be the reason a download fails."""
    questionnaires, calls = cache

    def boom(q_id):
        raise RuntimeError("ORA-00942")

    monkeypatch.setattr(questionnaires, "_export_version", boom)
    blob, name = questionnaires.export(41)
    assert blob and name.endswith(".xlsx")
    assert calls["render"] == 1


def test_missing_questionnaire_still_raises(cache, monkeypatch):
    from app.errors import NotFound
    questionnaires, _ = cache

    def missing(q_id):
        raise NotFound("Questionnaire 999 not found.")

    monkeypatch.setattr(questionnaires, "_export_version", missing)
    with pytest.raises(NotFound):
        questionnaires.export(999)


def test_fill_drops_the_cached_render():
    import inspect
    from app import questionnaires
    src = inspect.getsource(questionnaires.fill)
    assert "_EXPORT_CACHE.pop(q_id, None)" in src, (
        "a refill must not be servable from the pre-fill render"
    )


def test_version_probe_reads_only_counters():
    """It runs on every download; it must not touch the workbook blob."""
    import inspect
    from app import questionnaires
    src = inspect.getsource(questionnaires._export_version)
    assert "get_blob" not in src
    assert "load_workbook" not in src
