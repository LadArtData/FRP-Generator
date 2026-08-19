"""Regression tests for questionnaire workbook export.

Nashua's workbook 5262 merged its response column. openpyxl exposes every cell
of a merged range except the top-left anchor as a read-only MergedCell, so the
export raised "'MergedCell' object attribute 'value' is read-only" and returned
a 500, losing all 96 answers over cell geometry.
"""

import io

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell

from app.questionnaires import _write_cell


def _sheet_with_merges(*ranges):
    workbook = Workbook()
    sheet = workbook.active
    for rng in ranges:
        sheet.merge_cells(rng)
    return workbook, sheet


def test_plain_cell_writes_normally():
    workbook, sheet = _sheet_with_merges()
    assert _write_cell(sheet, "C4", "Yes") is True
    assert sheet["C4"].value == "Yes"


def test_merged_anchor_writes_normally():
    workbook, sheet = _sheet_with_merges("C4:E4")
    assert _write_cell(sheet, "C4", "Standard") is True
    assert sheet["C4"].value == "Standard"


def test_merged_follower_redirects_to_anchor():
    """The bug. D4 is a MergedCell; the write must land on C4, not raise."""
    workbook, sheet = _sheet_with_merges("C4:E4")
    assert isinstance(sheet["D4"], MergedCell)
    assert _write_cell(sheet, "D4", "Modification") is True
    assert sheet["C4"].value == "Modification"


def test_vertical_merge_redirects_to_top_left():
    workbook, sheet = _sheet_with_merges("B2:B6")
    assert _write_cell(sheet, "B5", "Third Party") is True
    assert sheet["B2"].value == "Third Party"


def test_block_merge_redirects_to_top_left():
    workbook, sheet = _sheet_with_merges("B2:D6")
    assert _write_cell(sheet, "D6", "Future Release") is True
    assert sheet["B2"].value == "Future Release"


def test_bad_coordinate_returns_false_instead_of_raising():
    workbook, sheet = _sheet_with_merges()
    assert _write_cell(sheet, "not-a-cell", "Yes") is False


def test_merged_workbook_survives_round_trip():
    """End to end: a merged response column must save and reload cleanly."""
    workbook, sheet = _sheet_with_merges("C2:E2", "C3:E3")
    sheet["A2"] = "Requirement one"
    sheet["A3"] = "Requirement two"
    assert _write_cell(sheet, "D2", "Standard")
    assert _write_cell(sheet, "E3", "Configuration")

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    reloaded = load_workbook(buffer).active
    assert reloaded["C2"].value == "Standard"
    assert reloaded["C3"].value == "Configuration"


def test_export_never_raises_on_read_only_cell():
    """_write_cell reports failure rather than propagating AttributeError, so a
    single unplaceable cell cannot cost the operator the other 95 answers."""
    workbook, sheet = _sheet_with_merges()

    class Immovable:
        @property
        def value(self):
            return None

        @value.setter
        def value(self, _):
            raise AttributeError("read-only")

    sheet._cells[(9, 9)] = Immovable()
    with pytest.raises(AttributeError):
        sheet.cell(row=9, column=9).value = "x"
    assert _write_cell(sheet, "I9", "x") is False
