"""Tests for the logic that can be exercised without the database, Claude, or
LibreOffice. These are the parts where a silent defect would be worst: document
classification (the corpus gate), chunking, JSON extraction from model output,
questionnaire column detection and Excel round-trip, response-code matching,
format validation, and session tokens.

They run offline. Database, Claude, and PDF conversion are integration concerns
verified on first deploy against the real ADB."""
import io
import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Config validation runs at import for some modules, so set the minimum env first.
os.environ.setdefault("ORACLE_PASSWORD", "x")
os.environ.setdefault("ORACLE_DSN", "x")
os.environ.setdefault("GENAI_REGION", "us-chicago-1")
os.environ.setdefault("GENAI_MODEL_OCID", "ocid1.generativeaimodel.oc1.test.aaaa")
os.environ.setdefault("GENAI_COMPARTMENT_ID", "ocid1.compartment.oc1..test")
os.environ.setdefault("HARALD_SESSION_SECRET", "unit-test-secret-value")
os.environ.setdefault("HARALD_APPROVER_PASSPHRASE", "unit-test-pass")


# ---------------------------------------------------------------------------
# classifier: the gate that keeps the corpus clean
# ---------------------------------------------------------------------------
from app import classifier


class TestClassifier:
    def test_iteria_narrative(self):
        assert classifier.classify_path("iteria_StPetersburg_Proposal.docx") == \
            classifier.ITERIA_NARRATIVE
        assert classifier.classify_path("iteria Technical Response.docx") == \
            classifier.ITERIA_NARRATIVE

    def test_competitor_is_not_narrative(self):
        assert classifier.classify_path("CanAm_Response.docx") == classifier.COMPETITOR
        assert classifier.classify_path("Drivestream_Proposal.docx") == classifier.COMPETITOR
        assert classifier.classify_path("Mythics_Final.docx") == classifier.COMPETITOR

    def test_client_rfp(self):
        assert classifier.classify_path("City_RFP_2026.pdf") == classifier.CLIENT_RFP
        assert classifier.classify_path("Attachment C - Requirements.docx") == \
            classifier.CLIENT_RFP

    def test_admin_and_pricing_excluded(self):
        assert classifier.classify_path("W-9_Form.pdf") == classifier.ADMIN
        assert classifier.classify_path("Non-Collusion Affidavit.pdf") == classifier.ADMIN
        assert classifier.classify_path("Pricing_Worksheet.xlsx") == classifier.PRICING

    def test_junk_excluded(self):
        assert classifier.classify_path("thumbnail.m97mqk") == classifier.EXCLUDE
        assert classifier.classify_path(".DS_Store") == classifier.EXCLUDE

    def test_module_detection_counts_hits(self):
        assert classifier.module_of(
            "The general ledger and accounts payable must post journal entries.") == "FIN"
        assert classifier.module_of(
            "Payroll processing including garnishments and W-2 earnings codes.") == "PAYROLL"
        assert classifier.module_of("A brief general statement.") == "GENERAL"

    def test_section_detection(self):
        assert classifier.section_of("Executive Summary") == "exec_summary"
        assert classifier.section_of("3.2 Implementation Approach") == "methodology"
        assert classifier.section_of("Firm Qualifications and Experience") == "qualifications"


# ---------------------------------------------------------------------------
# chunking: section-aware splitting
# ---------------------------------------------------------------------------
from app import chunking


class TestChunking:
    def test_heading_detection(self):
        assert chunking._is_heading("EXECUTIVE SUMMARY", "")
        assert chunking._is_heading("Solution Overview", "Heading 1")
        assert chunking._is_heading("2.1 Financial Management", "")
        assert not chunking._is_heading(
            "This is an ordinary sentence of body text that runs on.", "Normal")

    def test_chunk_breaks_on_heading(self):
        blocks = [
            ("H", "FINANCIAL MANAGEMENT"),
            ("P", "iteria configures the general ledger for fund accounting. " * 3),
            ("H", "HUMAN RESOURCES"),
            ("P", "Core HR and position management are delivered as standard. " * 3),
        ]
        chunks = chunking.chunk(blocks)
        assert len(chunks) >= 2
        assert any("general ledger" in c["text"] for c in chunks)
        assert any("Core HR" in c["text"] for c in chunks)

    def test_chunk_respects_size(self):
        blocks = [("P", "Sentence number %d in a long run of body text. " % i) for i in range(200)]
        chunks = chunking.chunk(blocks)
        assert len(chunks) > 1
        assert all(c["token_count"] <= chunking.MAX_TOKENS * 2 for c in chunks)

    def test_no_chunk_starts_mid_sentence(self):
        """The defect this chunker replaces: chunk 3 of the anchor document
        ended one cover letter and began the next."""
        body = ("The system supports fund accounting. iteria has delivered this "
                "for St. Petersburg and for Ozaukee County. Version 2.5 is "
                "current. Contact brian.schell@iteria.us for details. ") * 12
        chunks = chunking.chunk([("H", "TECHNICAL APPROACH"), ("P", body)])
        assert len(chunks) > 1
        for c in chunks:
            first = c["text"].split(". ")[0]
            assert first[:1].isupper(), f"chunk starts mid-sentence: {c['text'][:70]!r}"

    def test_abbreviations_do_not_split_sentences(self):
        sentences = chunking.split_sentences(
            "iteria served St. Petersburg. Mr. Poceous signed it. "
            "The rate is 2.5 percent. Email brian.schell@iteria.us today.")
        assert len(sentences) == 4, sentences
        assert "St. Petersburg" in sentences[0]
        assert "2.5" in sentences[2]

    def test_running_headers_are_stripped(self):
        blocks = [("P", "City of Example RFP Response")] * 4 + \
                 [("P", "Fund accounting is delivered as standard configuration. " * 4)]
        chunks = chunking.chunk(blocks)
        assert not any("City of Example RFP Response" in c["text"] for c in chunks)

    def test_tag_source_is_recorded(self):
        blocks = [("H", "COST PROPOSAL"),
                  ("P", "Our cost proposal totals $450,000 with a not-to-exceed "
                        "ceiling. Payment terms follow the milestone schedule. " * 3)]
        chunks = chunking.chunk(blocks)
        assert chunks
        assert chunks[0]["section"] == "cost"
        assert chunks[0]["tag_source"] in ("body", "smoothed")


# ---------------------------------------------------------------------------
# llm.parse_json: recover JSON from imperfect model output
# ---------------------------------------------------------------------------
from app import llm
from app.errors import UpstreamError


class TestParseJson:
    def test_plain_object(self):
        assert llm.parse_json('{"a": 1}') == {"a": 1}

    def test_strips_fences(self):
        assert llm.parse_json('```json\n{"a": 1}\n```') == {"a": 1}
        assert llm.parse_json('```\n{"a": 1}\n```') == {"a": 1}

    def test_recovers_from_prose(self):
        assert llm.parse_json('Here is the result: {"a": 1} thanks') == {"a": 1}

    def test_array(self):
        assert llm.parse_json('[{"x": 1}, {"x": 2}]', expect=list) == [{"x": 1}, {"x": 2}]

    def test_recovers_array_from_prose(self):
        assert llm.parse_json('The requirements are: [1, 2, 3].', expect=list) == [1, 2, 3]

    def test_type_mismatch_raises(self):
        with pytest.raises(ValueError):
            llm.parse_json('{"a": 1}', expect=list)

    def test_garbage_raises(self):
        with pytest.raises(ValueError):
            llm.parse_json("no json here at all")


# ---------------------------------------------------------------------------
# generation._match_code: map a chosen code to the workbook's allowed values
# ---------------------------------------------------------------------------
from app import generation


class TestMatchCode:
    def test_exact(self):
        assert generation._match_code("Standard", ["Standard", "Configuration"]) == "Standard"

    def test_case_insensitive(self):
        assert generation._match_code("standard", ["Standard", "Configuration"]) == "Standard"

    def test_substring(self):
        assert generation._match_code(
            "Config", ["Standard", "Configuration", "Modification"]) == "Configuration"

    def test_unmatched_falls_back_to_conservative(self):
        # No match returns the last (most conservative) allowed code, not a guess.
        assert generation._match_code(
            "Nonsense", ["Standard", "Not Available"]) == "Not Available"

    def test_empty_allowed_returns_input(self):
        assert generation._match_code("Anything", []) == "Anything"


# ---------------------------------------------------------------------------
# formats._validate_page_order
# ---------------------------------------------------------------------------
from app import formats
from app.errors import ValidationFailed


class TestFormatValidation:
    def test_valid_order_fills_keys(self):
        import json
        raw = formats._validate_page_order([{"title": "Executive Summary", "source": "generated"}])
        parsed = json.loads(raw)
        assert parsed[0]["key"] == "executive_summary"
        assert parsed[0]["source"] == "generated"

    def test_missing_title_rejected(self):
        with pytest.raises(ValidationFailed):
            formats._validate_page_order([{"source": "generated"}])

    def test_bad_source_rejected(self):
        with pytest.raises(ValidationFailed):
            formats._validate_page_order([{"title": "X", "source": "nonsense"}])

    def test_empty_rejected(self):
        with pytest.raises(ValidationFailed):
            formats._validate_page_order([])


# ---------------------------------------------------------------------------
# auth: signed session tokens and role ranking
# ---------------------------------------------------------------------------
from app import auth
from app.errors import Forbidden, Unauthorized


class TestAuth:
    def test_round_trip(self):
        token = auth.issue_token("amanda", "reviewer")
        claims = auth.parse_token(token)
        assert claims["username"] == "amanda"
        assert claims["role"] == "reviewer"

    def test_tampered_rejected(self):
        token = auth.issue_token("amanda", "reviewer")
        body, _, sig = token.partition(".")
        with pytest.raises(Unauthorized):
            auth.parse_token(body + "." + "x" * len(sig))

    def test_missing_rejected(self):
        with pytest.raises(Unauthorized):
            auth.parse_token(None)

    def test_role_ranking(self):
        auth.require({"role": "approver"}, auth.REVIEWER)   # ok, no raise
        auth.require({"role": "reviewer"}, auth.REVIEWER)   # ok
        with pytest.raises(Forbidden):
            auth.require({"role": "contributor"}, auth.REVIEWER)

    def test_approver_gate(self):
        auth.require_approver({"role": "approver"})
        with pytest.raises(Forbidden):
            auth.require_approver({"role": "reviewer"})


# ---------------------------------------------------------------------------
# questionnaires: column detection and Excel round-trip (openpyxl, no DB)
# ---------------------------------------------------------------------------
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from app import questionnaires


def _workbook_with_dropdown():
    wb = Workbook()
    ws = wb.active
    ws.title = "Financials"
    ws.append(["Req #", "Requirement Description", "Vendor Response", "Comments"])
    ws.append(["FIN-001", "The system shall support fund accounting.", "", ""])
    ws.append(["FIN-002", "The system shall provide bank reconciliation.", "", ""])
    ws.append(["FIN-003", "Describe chart of accounts flexibility.", "", ""])
    dv = DataValidation(
        type="list",
        formula1='"Standard,Configuration,Modification,Not Available"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add("C2:C100")
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class TestQuestionnaireDetection:
    def test_detects_columns_and_dropdown(self):
        wb = load_workbook(io.BytesIO(_workbook_with_dropdown()), data_only=True)
        detected = questionnaires._detect(wb.active)
        assert detected is not None
        from openpyxl.utils import get_column_letter
        assert get_column_letter(detected["question_col"]) == "B"
        assert get_column_letter(detected["response_col"]) == "C"
        assert get_column_letter(detected["comment_col"]) == "D"
        assert detected["allowed_codes"] == \
            ["Standard", "Configuration", "Modification", "Not Available"]

    def test_response_column_from_validation_without_keyword(self):
        # Header that does not say "response", but the column carries the dropdown.
        wb = Workbook()
        ws = wb.active
        ws.append(["Item", "Capability", "Disposition", "Notes"])
        ws.append(["1", "Fund accounting capability description here.", "", ""])
        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add("C2:C50")
        detected = questionnaires._detect(ws)
        from openpyxl.utils import get_column_letter
        assert get_column_letter(detected["response_col"]) == "C"
        assert detected["allowed_codes"] == ["Yes", "No"]

    def test_no_question_column_returns_none(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Alpha", "Beta", "Gamma"])
        ws.append(["1", "2", "3"])
        assert questionnaires._detect(ws) is None

    def test_validation_range_claims_only_its_column(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Requirement", "Response", "Comment"])
        ws.append(["A requirement statement of sufficient length.", "", ""])
        dv = DataValidation(type="list", formula1='"A,B,C"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add("B2:B10")
        columns = questionnaires._list_validations(ws)
        from openpyxl.utils import column_index_from_string
        assert column_index_from_string("B") in columns
        assert column_index_from_string("A") not in columns
        assert column_index_from_string("C") not in columns


class TestExcelRoundTrip:
    def test_write_back_preserves_dropdown(self):
        original = _workbook_with_dropdown()
        wb = load_workbook(io.BytesIO(original))
        ws = wb.active
        ws["C2"] = "Standard"
        ws["D2"] = "Native fund accounting supports this."
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        reloaded = load_workbook(out)
        rws = reloaded.active
        assert rws["C2"].value == "Standard"
        assert rws["D2"].value.startswith("Native fund accounting")
        assert len(rws.data_validations.dataValidation) > 0

    def test_formula_referenced_list_resolves(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Reqs"
        ws.append(["Requirement", "Response"])
        ws.append(["A sufficiently long requirement statement here.", ""])
        lists = wb.create_sheet("Lists")
        for i, value in enumerate(["Standard", "Config", "Mod"], start=1):
            lists.cell(row=i, column=1, value=value)
        dv = DataValidation(type="list", formula1="=Lists!$A$1:$A$3", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add("B2:B10")
        columns = questionnaires._list_validations(ws)
        from openpyxl.utils import column_index_from_string
        assert columns.get(column_index_from_string("B")) == ["Standard", "Config", "Mod"]
